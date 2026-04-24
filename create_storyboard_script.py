"""
create_storyboard_script.py
---------------------------
This file creates a storyboard script Excel file for the AI course.

Think of it like a TV script — it shows:
- What the narrator SAYS (audio column)
- What appears ON SCREEN (ost column)
- What GRAPHICS to show (graphics column — empty for now)

The output file has one section per screen, each with:
- A screen label (screen_1, screen_2, etc.)
- A header row (title | audio | ost | graphics)
- A data row with the actual content

For quiz screens (Check Your Understanding):
- Audio = "-" (no narration needed, learner reads silently)
- OST = The full question with answer options
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Course Content ────────────────────────────────────────────────────────────
# Each dictionary represents one screen in the course.
# Fields:
#   screen_label — the screen identifier (screen_1, screen_2, etc.)
#   title        — the name of the screen
#   audio        — what the narrator says out loud
#   ost          — text displayed on screen (On Screen Text)
# Note: graphics is always empty — we have no images in this course

screens = [
    {
        "screen_label": "screen_1",
        "title": "introduction",
        "audio": "Welcome to this course on Artificial Intelligence. In this module, you will explore how AI works, where it is used, and why responsible development matters.",
        "ost": "Artificial Intelligence (AI) is a rapidly evolving field of computer science that seeks to replicate human intelligence through intelligent machines. It enables learning, reasoning, and problem-solving. AI is transforming healthcare, education, transportation, and finance, while raising critical ethical concerns around privacy, bias, and responsibility.",
    },
    {
        "screen_label": "screen_2",
        "title": "learning objectives",
        "audio": "By the end of this course, you will be able to define Artificial Intelligence, distinguish between Narrow and General AI, understand machine learning, explain deep learning and neural networks, and identify ethical concerns in AI development.",
        "ost": "1. Define Artificial Intelligence and its role in computer science\n2. Distinguish between Narrow AI and General AI\n3. Understand machine learning as the core subset of AI\n4. Explain deep learning and neural networks\n5. Identify ethical concerns and the need for responsible AI development",
    },
    {
        "screen_label": "screen_3",
        "title": "ai fundamentals",
        "audio": "Artificial Intelligence refers to computer systems that perform tasks requiring human intelligence. AI uses algorithms and data to learn, make decisions, and improve over time. Its applications span healthcare, finance, transportation, and education.",
        "ost": "1. AI refers to computer systems that perform tasks requiring human intelligence\n2. AI uses algorithms and data to learn, make decisions, and improve over time\n3. It encompasses many techniques including machine learning and deep learning\n4. AI goals range from automating routine tasks to complex human interaction\n5. Applications span healthcare, finance, transportation, and education",
    },
    {
        "screen_label": "screen_4",
        "title": "check your understanding",
        "audio": "-",  # No audio for quiz screens — learner reads silently
        "ost": "Q: What is the primary focus of Artificial Intelligence?\n\nA) Creating intelligent machines that think like humans\nB) Developing algorithms for data analysis\nC) Designing computer networks\nD) Improving user experience\n\nCorrect Answer: A",
    },
    {
        "screen_label": "screen_5",
        "title": "what is artificial intelligence?",
        "audio": "AI is a branch of computer science focused on building intelligent systems. Early AI systems were rule-based, while modern AI uses machine learning. Narrow AI is designed for specific tasks and is what we use today. General AI, the ultimate goal, can perform any intellectual human task.",
        "ost": "1. AI is a branch of computer science focused on building intelligent systems\n2. Early AI systems were rule-based; modern AI uses machine learning\n3. Modern AI can improve over time without being explicitly reprogrammed\n4. Narrow AI is designed for specific tasks - all current AI is Narrow AI\n5. General AI, the ultimate goal, can perform any intellectual human task",
    },
    {
        "screen_label": "screen_6",
        "title": "check your understanding",
        "audio": "-",  # No audio for quiz screens
        "ost": "Q: What type of AI is currently used in most systems today?\n\nA) Narrow AI\nB) General AI\nC) Superintelligence\nD) Artificial General Intelligence\n\nCorrect Answer: A",
    },
    {
        "screen_label": "screen_7",
        "title": "machine learning basics",
        "audio": "Machine learning enables computers to learn from experience using algorithms. Deep learning uses multi-layered neural networks to recognize complex patterns. Popular frameworks include TensorFlow, PyTorch, and Scikit-learn, powering applications like recommendation systems and voice assistants.",
        "ost": "1. Machine learning enables computers to learn from experience via algorithms\n2. Deep learning uses multi-layered neural networks to recognize complex patterns\n3. Popular frameworks include TensorFlow, PyTorch, and Scikit-learn\n4. Applications include recommendation systems, spam filters, and voice assistants\n5. Machine learning is a core subset powering most modern AI applications",
    },
    {
        "screen_label": "screen_8",
        "title": "check your understanding",
        "audio": "-",  # No audio for quiz screens
        "ost": "Q: What type of networks does deep learning use to recognize patterns?\n\nA) Single-layered networks\nB) Multi-layered neural networks\nC) Rule-based systems\nD) Decision trees\n\nCorrect Answer: B",
    },
    {
        "screen_label": "screen_9",
        "title": "summary",
        "audio": "In summary, AI has evolved from rule-based systems to machine learning and deep learning. While Narrow AI dominates today, General AI remains the long-term goal. Responsible development and ethical frameworks are essential to harness AI safely.",
        "ost": "AI focuses on building systems that mimic human intelligence through machine learning and deep learning. All existing AI is Narrow AI - designed for specific tasks - though General AI remains the long-term goal. Ethical concerns including privacy, bias, and job displacement highlight the need for responsible development.",
    },
]


# ── Style Settings ────────────────────────────────────────────────────────────
# Header row colours — light red background with dark red text
HEADER_BG = "F4CCCC"  # Light red/salmon background
HEADER_FG = "CC0000"  # Dark red text


# ── Style Helper Functions ────────────────────────────────────────────────────

def fill(color):
    """Create a solid background fill with the given hex colour."""
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def thin_border():
    """Create a thin light grey border around a cell."""
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


# ── Build the Excel File ──────────────────────────────────────────────────────

# Create a new blank workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Storyboard"
ws.sheet_view.showGridLines = False  # Hide grid lines for a cleaner look

# Set column widths to fit the content
ws.column_dimensions["A"].width = 25  # Title column
ws.column_dimensions["B"].width = 45  # Audio column
ws.column_dimensions["C"].width = 55  # OST column
ws.column_dimensions["D"].width = 18  # Graphics column

# Track which row we're currently writing to
current_row = 1

# Loop through each screen and write it to the Excel file
for screen in screens:

    # ── Row 1: Screen Label (e.g. "screen_1") ────────────────────────────────
    ws.row_dimensions[current_row].height = 18
    cell = ws.cell(row=current_row, column=2, value=screen["screen_label"])
    cell.font = Font(name="Arial", size=10, bold=True, color="333333")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    current_row += 1

    # ── Row 2: Header Row (title | audio | ost | graphics) ───────────────────
    ws.row_dimensions[current_row].height = 20
    headers = ["title", "audio", "ost(on screen text)", "graphics"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.fill = fill(HEADER_BG)
        cell.font = Font(name="Arial", size=10, bold=True, color=HEADER_FG)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = thin_border()
    current_row += 1

    # ── Row 3: Data Row (actual content) ─────────────────────────────────────
    ws.row_dimensions[current_row].height = 100
    data = [
        screen["title"],   # Column A: Screen title
        screen["audio"],   # Column B: Narrator audio script
        screen["ost"],     # Column C: On screen text
        "",                # Column D: Graphics (always empty)
    ]
    for col, value in enumerate(data, 1):
        cell = ws.cell(row=current_row, column=col, value=value)
        cell.font = Font(name="Arial", size=10, bold=(col == 1), color="1A1A2E")
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        cell.border = thin_border()
        if col == 4:
            # Graphics column — grey italic text (it's empty but styled)
            cell.font = Font(name="Arial", size=10, color="999999", italic=True)
    current_row += 1

    # ── Row 4: Blank Spacer Row ───────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 10
    current_row += 1  # Leave a small gap between each screen

# Save the file
wb.save("storyboard_script_final.xlsx")
print("✅ storyboard_script_final.xlsx created!")
