
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from e2e_test import build_storyboard

COLORS = {
    "introduction": "4472C4",  # Blue
    "objectives":   "FFD966",  # Yellow
    "screen":       "70AD47",  # Green
    "cyu":          "ED7D31",  # Orange
    "summary":      "9B59B6",  # Purple
}

def export_excel(storyboard: list, filename: str = "storyboard.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Storyboard"

    # Header row
    headers = ["#", "Type", "Title", "Content"]
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for i, screen in enumerate(storyboard, 1):
        screen_type = screen["type"]
        title = screen["title"]

        # Build content text
        if screen_type == "cyu":
            options = screen.get("options", {})
            content = f"Q: {screen.get('question', '')}\n"
            for k, v in options.items():
                content += f"{k}) {v}\n"
            content += f"Correct: {screen.get('correct_answer', '')}"
        else:
            content = screen.get("content", "")

        row = i + 1
        fill_color = COLORS.get(screen_type, "FFFFFF")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        font = Font(size=11)
        alignment = Alignment(vertical="top", wrap_text=True)

        values = [i, screen_type, title, content]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 70

    # Row heights
    ws.row_dimensions[1].height = 25
    for i in range(2, len(storyboard) + 2):
        ws.row_dimensions[i].height = 80

    wb.save(filename)
    print(f"✅ Excel file saved as {filename}")
    return filename

if __name__ == "__main__":
    print("=== Building Storyboard ===\n")
    storyboard = build_storyboard("sample.pdf")
    print("\n=== Exporting to Excel ===\n")
    export_excel(storyboard)
    print("\n✅ Done! Open storyboard.xlsx to check:")
    print("   • Row colours by type")
    print("   • Column widths")
    print("   • Text wrapping")
