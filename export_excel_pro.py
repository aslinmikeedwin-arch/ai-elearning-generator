import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Colour palette: Midnight Executive ──────────────────────────────────────
P = {
    "navy":       "1E2761",
    "mid_navy":   "2E3F8F",
    "ice":        "CADCFC",
    "ice_light":  "EEF3FF",
    "white":      "FFFFFF",
    "dark_text":  "1A1A2E",
    "muted":      "64748B",
    "green":      "2ECC71",
    "orange":     "E67E22",
    "red":        "E74C3C",
    "purple":     "9B59B6",
    "row_alt":    "F7F9FF",
}

TYPE_COLORS = {
    "introduction": ("1E2761", "CADCFC"),   # navy bg, ice text
    "objectives":   ("2E3F8F", "FFFFFF"),   # mid-navy bg, white text
    "screen":       ("1A6B3C", "FFFFFF"),   # dark green bg, white text
    "cyu":          ("B85A00", "FFFFFF"),   # dark orange bg, white text
    "summary":      ("6B2FA0", "FFFFFF"),   # purple bg, white text
}

TYPE_LABELS = {
    "introduction": "📖  INTRODUCTION",
    "objectives":   "🎯  OBJECTIVES",
    "screen":       "🖥️   CONTENT SCREEN",
    "cyu":          "❓  CHECK YOUR UNDERSTANDING",
    "summary":      "📝  SUMMARY",
}

def thin_border():
    s = Side(style="thin", color="CADCFC")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium", color="1E2761")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(size=11, color="FFFFFF", bold=True):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=11, color="1A1A2E", bold=False, italic=False):
    return Font(name="Arial", size=size, bold=bold, color=color, italic=italic)

def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def wrap(horizontal="left", vertical="top"):
    return Alignment(horizontal=horizontal, vertical=vertical,
                     wrap_text=True, indent=1)

def export_excel_pro(storyboard: list, filename: str, topic: str = "Course Storyboard"):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Cover ────────────────────────────────────────────────────────
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.sheet_view.showGridLines = False
    ws_cover.column_dimensions["A"].width = 2
    ws_cover.column_dimensions["B"].width = 60
    ws_cover.column_dimensions["C"].width = 30

    # Banner
    for row in range(1, 9):
        ws_cover.row_dimensions[row].height = 22
    ws_cover.merge_cells("B1:C8")
    cell = ws_cover["B1"]
    cell.value = f"eLearning Storyboard\n{topic}"
    cell.fill = fill(P["navy"])
    cell.font = Font(name="Arial", size=28, bold=True, color=P["white"])
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)

    # Meta rows
    meta = [
        ("Document Type", "eLearning Course Storyboard"),
        ("Topic", topic),
        ("Total Screens", str(len(storyboard))),
        ("Content Screens", str(sum(1 for s in storyboard if s["type"] == "screen"))),
        ("Quiz Questions", str(sum(1 for s in storyboard if s["type"] == "cyu"))),
        ("Status", "Draft — For Review"),
    ]
    for i, (label, value) in enumerate(meta, 10):
        ws_cover.row_dimensions[i].height = 22
        ws_cover.merge_cells(f"B{i}:B{i}")
        ws_cover.merge_cells(f"C{i}:C{i}")
        lbl = ws_cover[f"B{i}"]
        lbl.value = label
        lbl.fill = fill(P["mid_navy"])
        lbl.font = hdr_font(11)
        lbl.alignment = wrap("left", "center")
        lbl.border = thin_border()
        val = ws_cover[f"C{i}"]
        val.value = value
        val.fill = fill(P["ice_light"])
        val.font = body_font(11, bold=(label == "Status"))
        val.alignment = wrap("left", "center")
        val.border = thin_border()

    # Legend
    ws_cover.row_dimensions[18].height = 18
    leg_title = ws_cover["B18"]
    leg_title.value = "SCREEN TYPE LEGEND"
    leg_title.fill = fill(P["navy"])
    leg_title.font = hdr_font(10)
    leg_title.alignment = wrap("left", "center")
    ws_cover.merge_cells("B18:C18")

    legend_items = [
        ("introduction", "Introduction — Course opening context"),
        ("objectives",   "Objectives — Learning goals"),
        ("screen",       "Content Screen — Core learning material"),
        ("cyu",          "Check Your Understanding — Quiz question"),
        ("summary",      "Summary — Closing recap"),
    ]
    for i, (stype, desc) in enumerate(legend_items, 19):
        ws_cover.row_dimensions[i].height = 20
        bg, fg = TYPE_COLORS[stype]
        tag = ws_cover[f"B{i}"]
        tag.value = TYPE_LABELS[stype]
        tag.fill = fill(bg)
        tag.font = hdr_font(10, fg)
        tag.alignment = wrap("left", "center")
        tag.border = thin_border()
        desc_cell = ws_cover[f"C{i}"]
        desc_cell.value = desc
        desc_cell.fill = fill(P["ice_light"])
        desc_cell.font = body_font(10)
        desc_cell.alignment = wrap("left", "center")
        desc_cell.border = thin_border()

    # ── Sheet 2: Storyboard ───────────────────────────────────────────────────
    ws = wb.create_sheet("Storyboard")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Column config: (header, width)
    cols = [
        ("#",           5),
        ("Screen Type", 22),
        ("Title",       30),
        ("Content",     65),
        ("Notes",       20),
    ]
    for i, (_, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title banner
    ws.merge_cells("A1:E1")
    banner = ws["A1"]
    banner.value = f"eLearning Storyboard  —  {topic}"
    banner.fill = fill(P["navy"])
    banner.font = Font(name="Arial", size=16, bold=True, color=P["white"])
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Header row
    ws.row_dimensions[2].height = 28
    for i, (header, _) in enumerate(cols, 1):
        cell = ws.cell(row=2, column=i)
        cell.value = header
        cell.fill = fill(P["mid_navy"])
        cell.font = hdr_font(11)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = thin_border()

    # Data rows
    for idx, screen in enumerate(storyboard, 1):
        row = idx + 2
        stype = screen["type"]
        bg, fg = TYPE_COLORS.get(stype, (P["white"], P["dark_text"]))
        alt = P["row_alt"] if idx % 2 == 0 else P["white"]

        # Build content string
        if stype == "cyu":
            opts = screen.get("options", {})
            content = f"Q: {screen.get('question', '')}\n\n"
            for k, v in opts.items():
                marker = "✓" if k == screen.get("correct_answer", "") else " "
                content += f"  {k}) {v}  {marker}\n"
            content += f"\nCorrect Answer: {screen.get('correct_answer', '')}"
        else:
            content = screen.get("content", "")

        values = [
            idx,
            TYPE_LABELS.get(stype, stype),
            screen.get("title", ""),
            content,
            "",
        ]

        ws.row_dimensions[row].height = 90

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border()

            if col == 1:
                # # column — navy badge style
                cell.fill = fill(bg)
                cell.font = Font(name="Arial", size=13, bold=True, color=fg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 2:
                # Type badge
                cell.fill = fill(bg)
                cell.font = Font(name="Arial", size=10, bold=True, color=fg)
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
            elif col == 3:
                # Title
                cell.fill = fill(P["ice_light"])
                cell.font = body_font(11, bold=True)
                cell.alignment = wrap("left", "center")
            elif col == 4:
                # Content
                cell.fill = fill(alt)
                cell.font = body_font(11)
                cell.alignment = wrap("left", "top")
            else:
                # Notes
                cell.fill = fill(P["ice_light"])
                cell.font = body_font(10, color=P["muted"], italic=True)
                cell.alignment = wrap("left", "top")

    # ── Sheet 3: Summary Stats ────────────────────────────────────────────────
    ws_stats = wb.create_sheet("Summary")
    ws_stats.sheet_view.showGridLines = False
    ws_stats.column_dimensions["B"].width = 35
    ws_stats.column_dimensions["C"].width = 20

    ws_stats.merge_cells("B1:C1")
    stat_banner = ws_stats["B1"]
    stat_banner.value = "Course Summary Statistics"
    stat_banner.fill = fill(P["navy"])
    stat_banner.font = Font(name="Arial", size=16, bold=True, color=P["white"])
    stat_banner.alignment = Alignment(horizontal="center", vertical="center")
    ws_stats.row_dimensions[1].height = 36

    type_counts = {}
    for s in storyboard:
        type_counts[s["type"]] = type_counts.get(s["type"], 0) + 1

    stats = [
        ("Total Screens", f'=Storyboard!A{len(storyboard)+2}'),
        ("Introduction Screens", type_counts.get("introduction", 0)),
        ("Objective Screens", type_counts.get("objectives", 0)),
        ("Content Screens", type_counts.get("screen", 0)),
        ("Quiz Questions (CYU)", type_counts.get("cyu", 0)),
        ("Summary Screens", type_counts.get("summary", 0)),
    ]

    for i, (label, value) in enumerate(stats, 3):
        ws_stats.row_dimensions[i].height = 24
        lbl = ws_stats[f"B{i}"]
        lbl.value = label
        lbl.fill = fill(P["mid_navy"])
        lbl.font = hdr_font(11)
        lbl.alignment = wrap("left", "center")
        lbl.border = thin_border()
        val = ws_stats[f"C{i}"]
        val.value = value
        val.fill = fill(P["ice_light"])
        val.font = body_font(12, bold=True)
        val.alignment = Alignment(horizontal="center", vertical="center")
        val.border = thin_border()

    wb.save(filename)
    print(f"✅ Professional Excel saved: {filename}")
    return filename


if __name__ == "__main__":
    from e2e_test import build_storyboard

    jobs = [
        ("sample.pdf",  "Artificial Intelligence",       "pro_sample1.xlsx"),
        ("sample2.pdf", "Cybersecurity and Data Privacy", "pro_sample2.xlsx"),
        ("sample3.pdf", "Business and Digital Marketing", "pro_sample3.xlsx"),
    ]

    for pdf, topic, out in jobs:
        print(f"\n📄 Processing: {topic}")
        storyboard = build_storyboard(pdf)
        export_excel_pro(storyboard, out, topic)

    print("\n🎉 All 3 professional Excel files created!")
